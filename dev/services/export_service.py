"""
export_service.py — Servicio de exportación de reportes a PDF y Excel.

Arquitectura
------------
Capa de utilidades para generar archivos exportables desde los reportes.
No contiene lógica de negocio — recibe datos ya procesados desde ReporteService
y los formatea en bytes descargables por el navegador via rx.download().

Formatos soportados:
    - PDF: Usando reportlab con tablas estilizadas.
    - Excel: Usando openpyxl con formato profesional.

Todos los métodos retornan `bytes` listos para pasarse a `rx.download(data=..., filename=...)`.

Uso desde la capa State:
    from dev.services.export_service import ExportService

    pdf_bytes = ExportService.generate_pdf(titulo, subtitulo, headers, rows)
    return rx.download(data=pdf_bytes, filename="reporte.pdf")
"""

from __future__ import annotations

import io
import logging
from datetime import date

logger = logging.getLogger("dev.services.export")


class ExportService:

    @classmethod
    def generate_pdf(
        cls,
        titulo: str,
        subtitulo: str,
        headers: list[str],
        rows: list[list[str]],
    ) -> bytes:
        """
        Genera un PDF en memoria y retorna los bytes.

        El PDF incluye:
            - Encabezado con título, subtítulo y fecha.
            - Tabla con bordes y colores alternos.
            - Pie con total de registros.

        Returns:
            bytes del archivo PDF generado.
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

        buffer = io.BytesIO()
        page_size = landscape(A4) if len(headers) > 5 else A4

        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(titulo, styles["Title"]))
        elements.append(Paragraph(subtitulo, styles["Normal"]))
        elements.append(
            Paragraph(
                f"Generado: {date.today().strftime('%d/%m/%Y')}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 0.5 * cm))

        if not rows:
            elements.append(Paragraph("No hay datos para este reporte.", styles["Normal"]))
        else:
            table_data = [headers] + rows
            usable_width = page_size[0] - doc.leftMargin - doc.rightMargin
            num_cols = len(table_data[0]) if table_data else 1
            col_widths = [usable_width / num_cols] * num_cols

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 0.5 * cm))
            elements.append(
                Paragraph(f"Total registros: {len(rows)}", styles["Normal"])
            )

        doc.build(elements)
        pdf_bytes = buffer.getvalue()
        buffer.close()

        logger.info("PDF generado en memoria (%s bytes, %s registros)", len(pdf_bytes), len(rows))
        return pdf_bytes

    @classmethod
    def generate_excel(
        cls,
        titulo: str,
        headers: list[str],
        rows: list[list[str]],
    ) -> bytes:
        """
        Genera un archivo Excel (.xlsx) en memoria y retorna los bytes.

        El Excel incluye:
            - Fila de encabezado estilizada.
            - Bordes, alineación y auto-width.
            - Filtro automático.

        Returns:
            bytes del archivo Excel generado.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]

        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, len(rows) + 2)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

        if rows:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

        buffer = io.BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()
        buffer.close()

        logger.info("Excel generado en memoria (%s bytes, %s registros)", len(xlsx_bytes), len(rows))
        return xlsx_bytes
